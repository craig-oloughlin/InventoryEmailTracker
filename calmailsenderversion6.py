#intending this program to be a standalone mailer that will send a weekly
#list of out of cal or nearly out of cal items picked out of a .xlsx file.

import openpyxl
from datetime import *
import win32com.client as win32
import time

file = 'K:\\[samplepath]\\Test_Field_Calibration-InventoryControl.xlsx'
archive_file_path = 'K:\\[samplepath]\\InventoryControl Archived\\' #needs a \\ at the end of the path
mail_to = "sample@sample.com"
subject_line = "Monthly Calibration Due List"
runner = ["<i>This is an automated message generated weekly by a Python script. Reply with ""calibration force update"" \
            in subject line to receive an immediately updated list.</i><br><br>"]

#print("Automated Calibration Status Updater.\n") #debugging and logging stuff
#print("%s: Begin Monitoring" % str(datetime.today())) 

def add_to_list(list_to_add, sheet, row, loc_option=False): #function: send a list to add an item to and the sheet and row the item is on.
    current_wb = wb.get_sheet_by_name(sheets[sheet])
    list_to_add.append("<tr><td width=""5%%"" align=""left"" valign=""middle"">%s</td>" % str(current_wb.cell(row=row, column=1).value))
    list_to_add.append("<td width=""45%%"" align=""left"" valign=""middle"">%s</td>" % str(current_wb.cell(row=row, column=2).value))
    list_to_add.append("<td width=""20%%"" align=""left"" valign=""middle"">SN: %s</td>" % str(current_wb.cell(row=row, column=3).value))
    if loc_option != True:
        date = current_wb.cell(row=row, column=7).value.date()
        list_to_add.append("<td><b>Cal Due: " + str(date) + '</b></td></tr>')
    else:
        list_to_add.append("<td><b>Location: " + str(location) + '</b></td></tr>')
        

def seal_list(list_to_seal):
    if len(list_to_seal) < 2:
        list_to_seal.append("</table>--None--<br><br>")
    else:
        list_to_seal.append('</table><br><br>')


def send_email(output):
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = mail_to
    mail.Subject = subject_line
    mail.HTMLBody = output
    mail.Send()

def send_email_to(send_to, message):
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = send_to
    mail.Subject = "Dog"
    mail.HTMLBody = message
    mail.Send()
    

running = True
sent = False

while running:
    
    outlookread = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = outlookread.GetDefaultFolder(6)
    messages = inbox.Items
    message = messages.GetLast()
    #print("%s: Condition Check" % str(datetime.today()))
    
    if (date.weekday(date.today()) == 6 and not sent) or "calibration force update" in str(message.Subject):

        message.Delete()

        due_for_cal = ["<b><u>The following are currently out of calibration:</u></b><br><br><table style=""width:100%"">"]
        due_for_cal_one_month = ["<b><u>The following are due for calibration within the next month:</u></b><br><br><table style=""width:100%"">"]
        due_for_cal_two_months = ["<b><u>The following are due for calibration within the next two months:</u></b><br><br><table style=""width:100%"">"]
        out_for_cal = ["<b><u>The following are currently out for calibration:</u></b><br><br><table style=""width:100%"">"]
        out_in_field = ["<u><b>The following are currently out in the field:</u></b><br><br><table style=""width:100%"">"]
    
        try:
            wb = openpyxl.load_workbook(file, data_only=True)              #read excel file into memory
            #print("%s: Workbook Open - Success" % str(datetime.today())) 
        except IOError as ioerror:
            send_email("An error has occurred while trying to access the Inventory Control Excel sheet on the network.\
                            Please verify that the file has not been moved or renamed.")
            #print("%s: Error - Could not access workbook." % str(datetime.today())) 
    
        sheets = wb.get_sheet_names()
        for i in range(0, len(sheets), 1):                                 #iterate each sheet
            current_wb = wb.get_sheet_by_name(sheets[i])
            for j in range(1, current_wb.max_row, 1):                      #iterate each row
                caldue = current_wb.cell(row=j, column=7).value
                location = current_wb.cell(row=j, column=8).value
        
                if type(caldue) == datetime and location != "Calibration": #series of checks on each rows data values
            
                    if  caldue <= datetime.today():
                        add_to_list(due_for_cal, i, j)
                
                    if caldue > datetime.today() and caldue <= (datetime.today() + timedelta(31)):
                        add_to_list(due_for_cal_one_month, i, j)

                    if caldue > (datetime.today() + timedelta(31)) and caldue <= (datetime.today() + timedelta(62)):
                        add_to_list(due_for_cal_two_months, i, j)

                if location == "Calibration" and type(caldue) == datetime:
                    add_to_list(out_for_cal, i, j)

                if location != "Calibration" and "Base-" not in str(location) and type(caldue) == datetime:
                    add_to_list(out_in_field, i, j, loc_option=True)
                    
        #finish and email the lists

        seal_list(due_for_cal)
        seal_list(due_for_cal_one_month)
        seal_list(due_for_cal_two_months)
        seal_list(out_for_cal)
        seal_list(out_in_field)
        output_list = due_for_cal + due_for_cal_one_month + due_for_cal_two_months + out_for_cal + out_in_field + runner
        output = ''.join(output_list)
        
        try:
            send_email(output)
            sent = True
            #print("%s: Email Send - Success" % str(datetime.today()))
        except:
            pass#print("%s: Error - Email could not be sent" % str(datetime.today()))

        #clear the lists for the next run

        del due_for_cal[:]
        del due_for_cal_one_month[:]
        del due_for_cal_two_months[:]
        del out_for_cal [:]
        del out_in_field [:]

        #save a copy for the archive

        wb2 = openpyxl.load_workbook(file)   #need to reload the file as the previous load was data only, no formula
        try:
            wb2.save('%s%s.xlsx' % (archive_file_path, str(date.today())))
        except IOError as ioerror:
            send_email("Notice: There was an error while trying to save an archive copy of the inventory control document.\
                        Please check the archive path.")
            #print("%s: Error - Could not save an archive copy" % str(datetime.today()))
             
    if date.weekday(date.today()) != 6:

        sent = False

    if "bark" == str(message.Subject):

        message.Delete()
        send_email_to("sample@sample.com", "woof")
       
    time.sleep(10)  # check the condition every 10 seconds. Hopefully
